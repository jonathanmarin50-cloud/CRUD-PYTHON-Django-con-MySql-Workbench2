"""
Vistas CRUD de la aplicación de gestión de pedidos.
Aquí se encuentra toda la 'lógica' del negocio. Las vistas reciben 
las peticiones del usuario, consultan la base de datos (MySQL) 
y deciden qué página de HTML mostrar.
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache

from .models import Cliente, Producto, Pedido, DetallePedido
from .forms import ClienteForm, ProductoForm, PedidoForm, DetallePedidoForm

# --- Librerías externas para exportar datos ---
import openpyxl  # Para Excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF  # Para PDF
from datetime import date
import io # Para manejo de flujos de datos binarios

# ============================================================
#  DASHBOARD (PÁGINA PRINCIPAL)
# ============================================================
@login_required # Obliga a que el usuario haya iniciado sesión.
@never_cache # Asegura que al dar 'atrás' no se vean datos privados si cerró sesión.
def dashboard(request):
    """
    Controlador del resumen general.
    Consulta las estadísticas de todas las tablas para mostrarlas en las tarjetas.
    """
    context = {
        'total_clientes': Cliente.objects.count(),
        'total_productos': Producto.objects.count(),
        'total_pedidos': Pedido.objects.count(),
        'pedidos_pendientes': Pedido.objects.filter(estado='Pendiente').count(),
        'pedidos_enviados': Pedido.objects.filter(estado='Enviado').count(),
        'pedidos_entregados': Pedido.objects.filter(estado='Entregado').count(),
        'ultimos_pedidos': Pedido.objects.select_related('cliente').order_by('-creado_en')[:5],
    }
    return render(request, 'gestion/dashboard.html', context)


# ============================================================
#  CLIENTES - OPERACIONES CRUD
# ============================================================

@login_required
@never_cache
def cliente_list(request):
    """Lista todos los clientes con buscador y paginación."""
    query = request.GET.get('q', '')
    clientes = Cliente.objects.all()
    if query:
        clientes = clientes.filter(
            Q(nombre__icontains=query) | Q(correo__icontains=query)
        )
    paginator = Paginator(clientes, 5)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'gestion/cliente_list.html', {'page_obj': page_obj, 'query': query})

@login_required
def cliente_create(request):
    """Formulario para crear un cliente nuevo."""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Cliente creado correctamente.")
            return redirect('cliente_list')
    else:
        form = ClienteForm()
    return render(request, 'gestion/cliente_form.html', {'form': form, 'titulo': 'Nuevo Cliente'})

@login_required
def cliente_update(request, pk):
    """Formulario para editar un cliente existente."""
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Cliente actualizado.")
            return redirect('cliente_list')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'gestion/cliente_form.html', {'form': form, 'titulo': 'Editar Cliente'})

@login_required
def cliente_delete(request, pk):
    """Borrar un cliente."""
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, "🗑️ Cliente eliminado.")
        return redirect('cliente_list')
    return render(request, 'gestion/confirm_delete.html', {'objeto': cliente})


# ============================================================
#  PRODUCTOS - OPERACIONES CRUD
# ============================================================

@login_required
def producto_list(request):
    """Lista de productos."""
    query = request.GET.get('q', '')
    productos = Producto.objects.all()
    if query:
        productos = productos.filter(nombre__icontains=query)
    paginator = Paginator(productos, 5)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'gestion/producto_list.html', {'page_obj': page_obj, 'query': query})

@login_required
def producto_create(request):
    """Crear producto."""
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Producto agregado.")
            return redirect('producto_list')
    else:
        form = ProductoForm()
    return render(request, 'gestion/producto_form.html', {'form': form, 'titulo': 'Nuevo Producto'})

@login_required
def producto_update(request, pk):
    """Editar producto."""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Producto actualizado.")
            return redirect('producto_list')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'gestion/producto_form.html', {'form': form, 'titulo': 'Editar Producto'})

@login_required
def producto_delete(request, pk):
    """Eliminar producto."""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.delete()
        messages.success(request, "🗑️ Producto eliminado.")
        return redirect('producto_list')
    return render(request, 'gestion/confirm_delete.html', {'objeto': producto})


# ============================================================
#  PEDIDOS - OPERACIONES CRUD
# ============================================================

@login_required
def pedido_list(request):
    """Lista de pedidos con buscador y filtro por estado."""
    query = request.GET.get('q', '')
    estado_filter = request.GET.get('estado', '')
    
    # Empezamos con todos los pedidos, ordenados por fecha (más recientes primero)
    pedidos = Pedido.objects.select_related('cliente').all()

    # Filtro por texto (Buscador por nombre de cliente)
    if query:
        pedidos = pedidos.filter(cliente__nombre__icontains=query)

    # Filtro por estado del pedido
    if estado_filter:
        pedidos = pedidos.filter(estado=estado_filter)

    total = pedidos.count()
    paginator = Paginator(pedidos, 5)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Pasamos las opciones de estado para el select del filtro (vienen del modelo)
    estados = Pedido.ESTADO_CHOICES

    return render(request, 'gestion/pedido_list.html', {
        'page_obj': page_obj, 
        'query': query, 
        'estado_filter': estado_filter,
        'estados': estados,
        'total': total
    })

@login_required
def pedido_create(request):
    """Crear una nueva cabecera de pedido."""
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            pedido = form.save()
            messages.success(request, "✅ Pedido iniciado. Ahora agrega productos.")
            return redirect('pedido_detail', pk=pedido.pk)
    else:
        form = PedidoForm()
    return render(request, 'gestion/pedido_form.html', {'form': form, 'titulo': 'Iniciar Pedido'})

@login_required
def pedido_update(request, pk):
    """Editar estado o cliente del pedido."""
    pedido = get_object_or_404(Pedido, pk=pk)
    if request.method == 'POST':
        form = PedidoForm(request.POST, instance=pedido)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Pedido actualizado.")
            return redirect('pedido_list')
    else:
        form = PedidoForm(instance=pedido)
    return render(request, 'gestion/pedido_form.html', {'form': form, 'titulo': 'Editar Pedido'})

@login_required
def pedido_delete(request, pk):
    """Borrar pedido completo."""
    pedido = get_object_or_404(Pedido, pk=pk)
    if request.method == 'POST':
        pedido.delete()
        messages.success(request, "🗑️ Pedido eliminado.")
        return redirect('pedido_list')
    return render(request, 'gestion/confirm_delete.html', {'objeto': pedido})

@login_required
def pedido_detail(request, pk):
    """Muestra el pedido y permite agregar productos (detalles)."""
    pedido = get_object_or_404(Pedido, pk=pk)
    detalles = pedido.detalles.select_related('producto').all()
    if request.method == 'POST':
        form = DetallePedidoForm(request.POST)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.pedido = pedido
            detalle.save()
            messages.success(request, "✅ Producto agregado al detalle.")
            return redirect('pedido_detail', pk=pk)
    else:
        form = DetallePedidoForm()
    return render(request, 'gestion/pedido_detail.html', {
        'pedido': pedido, 'detalles': detalles, 'form': form, 'total': pedido.total()
    })

@login_required
def detalle_delete(request, pk):
    """Eliminar un producto específico del pedido."""
    detalle = get_object_or_404(DetallePedido, pk=pk)
    pk_pedido = detalle.pedido.pk
    detalle.delete()
    messages.success(request, "🗑️ Producto quitado del pedido.")
    return redirect('pedido_detail', pk=pk_pedido)


# ============================================================
#  EXPORTACIÓN Y REGISTRO
# ============================================================

@login_required
def export_pedidos_excel(request):
    """Exportar pedidos a Excel."""
    pedidos = Pedido.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    ws.append(['ID', 'Fecha', 'Estado', 'Total'])
    for p in pedidos:
        ws.append([p.id, str(p.fecha), p.estado, float(p.total())])
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="pedidos_{date.today()}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_clientes_excel(request):
    """Exportar todos los clientes a Excel."""
    clientes = Cliente.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    # Encabezados
    ws.append(['ID', 'Nombre', 'Cédula/RIT', 'Correo', 'Teléfono', 'Dirección'])
    # Datos
    for c in clientes:
        ws.append([c.id, c.nombre, c.cedula, c.correo, c.telefono, c.direccion])
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="clientes_{date.today()}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_productos_excel(request):
    """Exportar todos los productos a Excel."""
    productos = Producto.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    # Encabezados
    ws.append(['ID', 'Nombre', 'Precio', 'Stock'])
    # Datos
    for p in productos:
        ws.append([p.id, p.nombre, float(p.precio), p.stock])
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="productos_{date.today()}.xlsx"'
    wb.save(response)
    return response

# Función auxiliar global para limpiar texto de caracteres no soportados por fuentes estándar (latin-1)
def clean_pdf_text(text):
    if not text: return ""
    return str(text).encode('latin-1', 'replace').decode('latin-1')

# --- Clase personalizada para PDF Premium ---
class PDFReporte(FPDF):
    def header(self):
        self.set_font("helvetica", "B", 15)
        self.set_text_color(33, 37, 41) 
        # Limpiamos el título para evitar errores con la 'Ó'
        titulo = clean_pdf_text("SISTEMA DE GESTION DE PEDIDOS")
        self.cell(0, 10, titulo, ln=True, align="C")
        self.set_font("helvetica", "I", 10)
        info = clean_pdf_text(f"Reporte Generado el: {date.today().strftime('%d/%m/%Y')}")
        self.cell(0, 10, info, ln=True, align="C")
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.set_text_color(128)
        pie = clean_pdf_text(f"Pagina {self.page_no()}/{{nb}}")
        self.cell(0, 10, pie, align="C")

@login_required
def export_pedidos_pdf(request):
    """Exportar pedidos a PDF con diseño profesional y corregido."""
    try:
        pedidos = Pedido.objects.select_related('cliente').all()
        
        pdf = PDFReporte()
        pdf.alias_nb_pages()
        pdf.add_page()
        
        # Título de la tabla
        pdf.set_font("helvetica", "B", 14)
        pdf.set_fill_color(52, 58, 64) 
        pdf.set_text_color(255, 255, 255) 
        pdf.cell(190, 10, clean_pdf_text(" LISTADO GENERAL DE PEDIDOS"), ln=True, align="L", fill=True)
        pdf.ln(2)

        # Cabecera de la tabla
        pdf.set_font("helvetica", "B", 10)
        pdf.set_fill_color(248, 249, 250) 
        pdf.set_text_color(0)
        
        pdf.cell(15, 8, "ID", border=1, align="C", fill=True)
        pdf.cell(85, 8, "Cliente", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Fecha", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Estado", border=1, align="C", fill=True)
        pdf.cell(30, 8, "Total", border=1, align="C", fill=True)
        pdf.ln()

        # Cuerpo de la tabla
        pdf.set_font("helvetica", "", 9)
        total_general = 0
        
        if not pedidos.exists():
            pdf.cell(190, 10, clean_pdf_text("No hay pedidos registrados."), border=1, ln=True, align="C")
        else:
            for p in pedidos:
                t = p.total()
                total_general += t
                
                pdf.cell(15, 7, f"{p.id}", border=1, align="C")
                
                nombre = clean_pdf_text(p.cliente.nombre)
                nombre_display = (nombre[:40] + '..') if len(nombre) > 40 else nombre
                
                pdf.cell(85, 7, f" {nombre_display}", border=1)
                pdf.cell(30, 7, f"{p.fecha.strftime('%d/%m/%Y')}", border=1, align="C")
                pdf.cell(30, 7, clean_pdf_text(p.estado), border=1, align="C")
                pdf.cell(30, 7, f"${t:,.2f}", border=1, align="R")
                pdf.ln()

        # Fila de Resumen
        pdf.set_font("helvetica", "B", 10)
        pdf.set_fill_color(233, 236, 239)
        pdf.cell(160, 8, clean_pdf_text("TOTAL VENTAS ACUMULADAS"), border=1, align="R", fill=True)
        pdf.cell(30, 8, f"${total_general:,.2f}", border=1, align="R", fill=True)

        # Entrega de datos binarios
        pdf_bytes = bytes(pdf.output())
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_pedidos.pdf"'
        response['Content-Length'] = len(pdf_bytes)
        return response

    except Exception as e:
        # Si falla, devolvemos el error como texto para saber qué pasó exactamente
        return HttpResponse(f"Error critico al generar PDF: {str(e)}", content_type="text/plain", status=500)

from django.contrib.auth.forms import UserCreationForm
def registro_usuario(request):
    """Registro de nuevos usuarios."""
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "🎉 Usuario creado.")
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'registration/registro.html', {'form': form})

# ============================================================
#  API REST (ENDPOINTS PARA INTEGRACIONES)
# ============================================================

def api_cliente_list(request):
    """Retorna la lista de clientes en formato JSON."""
    clientes = Cliente.objects.all().values('id', 'nombre', 'correo', 'telefono', 'direccion')
    return JsonResponse(list(clientes), safe=False)

def api_producto_list(request):
    """Retorna la lista de productos en formato JSON."""
    productos = Producto.objects.all().values('id', 'nombre', 'precio', 'stock')
    # Convertimos los Decimal de precio a Float para que JSON lo soporte
    lista_productos = []
    for p in productos:
        p['precio'] = float(p['precio'])
        lista_productos.append(p)
    return JsonResponse(lista_productos, safe=False)

@login_required
def api_docs(request):
    """Muestra la documentación estilo Swagger de la API."""
    return render(request, 'gestion/api_docs.html')

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def api_pedido_delete(request, pk):
    """Endpoint real para borrar pedidos vía API."""
    if request.method == 'DELETE':
        try:
            pedido = Pedido.objects.get(pk=pk)
            pedido.delete()
            return JsonResponse({'status': 'success', 'message': f'Pedido {pk} eliminado correctamente'}, status=200)
        except Pedido.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Pedido no encontrado'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
